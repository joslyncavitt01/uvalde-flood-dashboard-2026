"""One-off/periodic manual load of four ShelterLuv "flood week" custom reports --
diagnostic tests, vaccines, physical exams, and surgeries -- each an event-level log
(multiple rows per animal), unlike the single-row-per-animal profile snapshot. NOT
scoped to flood-attributable animals specifically; these are org-wide reports for the
date window, filtered down to flood animals later in fetch_data.py via a join.

The surgeries report exists separately from the data team's own CompletedSurgeries
table because that table's sync stopped updating before 2026-06-17 -- well before the
flood -- so it can't be relied on for anything in the flood window.

Each animal's "Attributes" tag list shows up in all four reports and is confirmed
consistent wherever it appears (checked directly), so it's read here but not stored
separately -- fetch_data.py pulls it live from whichever table has it for a given
animal.

Refresh model: same as backfill_animal_profiles.py -- manual, whenever Joslyn wants
this brought current. Each run fully replaces the destination table's contents
(WRITE_TRUNCATE load job, not DML), since these are periodic whole-window snapshots,
not incremental logs -- there's no stable per-row key to dedupe against across runs.

Usage: python3 backfill_animal_medical.py /path/to/floodweekdiagnostictests.xlsx \
    /path/to/floodweekvaccines.xlsx /path/to/floodweekphysicalexams.xlsx \
    /path/to/floodweeksurgeries.xlsx
"""
import sys
import openpyxl
from google.cloud import bigquery

PROJECT = "apa-data-410213"
DATASET = "shelterluv"

DIAGNOSTICS_MAP = {
    "Animal ID": "AnimalID",
    "Name": "Name",
    "Species": "Species",
    "Primary Breed": "PrimaryBreed",
    "Current Location": "CurrentLocation",
    "Attributes": "Attributes",
    "Test Date": "TestDate",
    "Test Status": "TestStatus",
    "Test Name": "TestName",
    "Test Product": "TestProduct",
    "Test By": "TestBy",
    "Test Notes": "TestNotes",
    "Result Name": "ResultName",
    "Result": "Result",
}

VACCINES_MAP = {
    "Animal ID": "AnimalID",
    "Name": "Name",
    "Species": "Species",
    "Current Location": "CurrentLocation",
    "Attributes": "Attributes",
    "Date Completed": "DateCompleted",
    "Vaccine Product": "VaccineProduct",
    "Lot #": "LotNumber",
    "Vaccinated By": "VaccinatedBy",
    "Rabies Tag Number": "RabiesTagNumber",
    "Supervising Veterinarian": "SupervisingVeterinarian",
}

SURGERIES_MAP = {
    "Animal ID": "AnimalID",
    "Name": "Name",
    "Species": "Species",
    "Primary Breed": "PrimaryBreed",
    "Sex": "Sex",
    "Age (Y/M/D)": "AgeYMD",
    "Altered In Care": "AlteredInCare",
    "Altered Before Arrival": "AlteredBeforeArrival",
    "Current Location": "CurrentLocation",
    "Current Status": "CurrentStatus",
    "Attributes": "Attributes",
    "Current Weight": "CurrentWeight",
    "Date Completed": "DateCompleted",
    "Procedure/Surgery Type": "SurgeryType",
    "Surgeon": "Surgeon",
    "Clinic": "Clinic",
    "Memo": "Memo",
}

EXAMS_MAP = {
    "Animal ID": "AnimalID",
    "Name": "Name",
    "Species": "Species",
    "Primary Breed": "PrimaryBreed",
    "Secondary Breed": "SecondaryBreed",
    "Sex": "Sex",
    "Altered": "Altered",
    "Current Location": "CurrentLocation",
    "Current Status": "CurrentStatus",
    "Attributes": "Attributes",
    "Date Completed": "DateCompleted",
    "Vet or Tech Exam": "VetOrTechExam",
    "Type": "Type",
    "Exam Reason": "ExamReason",
    "Subjective": "Subjective",
    "Objective": "Objective",
    "Assessment": "Assessment",
    "Plan": "Plan",
    "New Diagnoses": "NewDiagnoses",
    "Performed By": "PerformedBy",
}

TARGETS = {
    "diagnostic": ("DiagnosticTestsJoslyn", DIAGNOSTICS_MAP),
    "vaccine": ("VaccinesJoslyn", VACCINES_MAP),
    "exam": ("PhysicalExamsJoslyn", EXAMS_MAP),
    "surgery": ("SurgeriesJoslyn", SURGERIES_MAP),
}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "—", "-"):
        return None
    return s


def detect_kind(headers):
    if "Test Name" in headers:
        return "diagnostic"
    if "Vaccine Product" in headers:
        return "vaccine"
    if "Exam Reason" in headers:
        return "exam"
    if "Procedure/Surgery Type" in headers:
        return "surgery"
    raise ValueError(f"Couldn't identify report type from headers: {headers[:5]}...")


def load_file(client, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}

    kind = detect_kind(headers)
    table_name, field_map = TARGETS[kind]

    rows_out = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for report_col, bq_col in field_map.items():
            idx = col_idx.get(report_col)
            row[bq_col] = clean(raw[idx]) if idx is not None else None
        if row.get("AnimalID"):
            row["AnimalID"] = row["AnimalID"].replace("APA-A-", "")
            rows_out.append(row)

    print(f"{path.split('/')[-1]}: parsed {len(rows_out)} rows -> {table_name} ({kind})")

    table_ref = f"{PROJECT}.{DATASET}.{table_name}"
    job_config = bigquery.LoadJobConfig(
        write_disposition="WRITE_TRUNCATE",
        autodetect=True,
        source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
    )
    job = client.load_table_from_json(rows_out, table_ref, job_config=job_config)
    job.result()
    print(f"  Loaded {len(rows_out)} rows into {table_ref} (full replace).")


def run(paths):
    client = bigquery.Client(project=PROJECT)
    for path in paths:
        load_file(client, path)


if __name__ == "__main__":
    run(sys.argv[1:])
