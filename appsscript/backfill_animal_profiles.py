"""One-off/periodic manual load of apa-grants-manager.shelterluv.animal_profile_snapshot
from a "current animal snapshot" custom report downloaded directly from ShelterLuv (NOT the
hourly flood update report -- this is the much richer, ~200-column export with photos,
breed/color detail, and dozens of memo fields).

Refresh model: manual, whenever Joslyn wants profile detail (photos, breed, memos) brought
current -- unlike the hourly status/location pipeline, this data is mostly slow-changing.
Re-running with a newer export simply overwrites each animal's row (keyed by AnimalID),
so it's safe to re-run as often as she likes.

Usage: python3 backfill_animal_profiles.py /path/to/custom-report-*.xlsx
"""
import sys
from datetime import datetime, timezone
import openpyxl
from google.cloud import bigquery

PROJECT = "apa-data-410213"
DATASET = "shelterluv"
TABLE = "AnimalProfileSnapshotJoslyn"

# Straightforward 1:1 field mappings (report column name -> BQ column name)
FIELD_MAP = {
    "Animal ID": "AnimalID",
    "Internal ID (API)": "InternalID",
    "Name": "Name",
    "Species": "Species",
    "Primary Breed": "PrimaryBreed",
    "Secondary Breed": "SecondaryBreed",
    "Sex": "Sex",
    "Age (Y/M/D)": "AgeYMD",
    "Age (Months)": "AgeMonths",
    "Age Group": "AgeGroup",
    "Birthdate": "Birthdate",
    "Primary Color": "PrimaryColor",
    "Secondary Color": "SecondaryColor",
    "Pattern": "Pattern",
    "Microchip Number": "MicrochipNumber",
    "Microchip Issuer": "MicrochipIssuer",
    "Adoption Category": "AdoptionCategory",
    "Behavior Category": "BehaviorCategory",
    "Medical Category": "MedicalCategory",
    "Volunteer Category": "VolunteerCategory",
    "Altered In Care": "AlteredInCare",
    "Altered Before Arrival": "AlteredBeforeArrival",
    "Current Location": "CurrentLocation",
    "Days In Custody": "DaysInCustody",
    "Days Onsite": "DaysOnsite",
    "Days Available": "DaysAvailable",
    "Current Status": "CurrentStatus",
    "Visible": "Visible",
    "Holdable": "Holdable",
    "Adoptable": "Adoptable",
    "Onsite": "Onsite",
    "In Custody": "InCustody",
    "Current Weight": "CurrentWeight",
    "Transfer From": "TransferFrom",
    "Original Source": "OriginalSource",
    "Adoption Fee Group": "AdoptionFeeGroup",
    "Foster Person Name": "FosterPersonName",
    "Foster Person City": "FosterPersonCity",
    "Foster Person State": "FosterPersonState",
    "Photo": "Photo",
    "Video": "Video",
    "Date Created": "DateCreated",
    "Intake Date": "IntakeDate",
}

KENNEL_CARD_COL = "Kennel Card / Website Memo"

# Everything from here to the end of the header row is a "memo"-type field (behavioral,
# medical, foster notes, training/behavior-program fields, etc.). Only the ones actually
# filled in for a given animal get kept, as a JSON blob -- there are ~130 of these and the
# overwhelming majority are blank for any given animal.
MEMO_START_COL = "Behavioral"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "—", "-"):
        return None
    return s


def run(path):
    import json as jsonlib

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}

    memo_start_idx = col_idx.get(MEMO_START_COL)
    memo_cols = headers[memo_start_idx:] if memo_start_idx is not None else []

    ingested_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    source_file = path.split("/")[-1]

    rows_out = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {"IngestedAt": ingested_at, "SourceFile": source_file}
        for report_col, bq_col in FIELD_MAP.items():
            idx = col_idx.get(report_col)
            row[bq_col] = clean(raw[idx]) if idx is not None else None
        if row.get("AnimalID"):
            row["AnimalID"] = row["AnimalID"].replace("APA-A-", "")

        kc_idx = col_idx.get(KENNEL_CARD_COL)
        row["KennelCardMemo"] = clean(raw[kc_idx]) if kc_idx is not None else None

        memos = {}
        for h in memo_cols:
            idx = col_idx.get(h)
            if idx is None:
                continue
            val = clean(raw[idx])
            if val:
                memos[h.strip()] = val
        row["MemosJSON"] = jsonlib.dumps(memos) if memos else None

        if row.get("AnimalID"):
            rows_out.append(row)

    # ShelterLuv's own report export can contain duplicate rows for the same animal (seen
    # directly in a 2026-07-30 export) -- dedupe here so a single load never writes more
    # than one row per AnimalID, regardless of what the source file contains.
    deduped = {}
    for row in rows_out:
        deduped[row["AnimalID"]] = row
    dupes_dropped = len(rows_out) - len(deduped)
    rows_out = list(deduped.values())

    print(f"Parsed {len(rows_out)} animal profiles ({len(memo_cols)} candidate memo columns scanned, {dupes_dropped} duplicate rows dropped).")

    client = bigquery.Client(project=PROJECT)
    table_ref = f"{PROJECT}.{DATASET}.{TABLE}"

    # Delete any existing rows for these animal IDs first (this report is a full current
    # snapshot per animal, not an incremental log like the hourly one -- we want the latest
    # profile to fully replace the old one, not accumulate duplicate rows per animal).
    ids = [r["AnimalID"] for r in rows_out]
    if ids:
        id_list = ", ".join(f"'{i}'" for i in ids)
        client.query(f"DELETE FROM `{table_ref}` WHERE AnimalID IN ({id_list})").result()

    chunk_size = 500
    loaded = 0
    had_errors = False
    for i in range(0, len(rows_out), chunk_size):
        chunk = rows_out[i:i + chunk_size]
        errors = client.insert_rows_json(table_ref, chunk, row_ids=[None] * len(chunk))
        if errors:
            had_errors = True
            print(f"Errors during insert (rows {i}-{i+len(chunk)}):")
            for e in errors[:5]:
                print(" ", e)
        else:
            loaded += len(chunk)
    if not had_errors:
        print(f"Loaded {loaded} animal profiles into {table_ref}.")
    else:
        print(f"Loaded {loaded} of {len(rows_out)} animal profiles into {table_ref} (some chunks had errors).")


if __name__ == "__main__":
    run(sys.argv[1])
