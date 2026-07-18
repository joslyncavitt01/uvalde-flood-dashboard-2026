"""One-off backfill of apa-data-410213.shelterluv.TempFloodJoslyn from a manually
downloaded ShelterLuv custom report, to give the email-supplement pipeline a running
start before the Apps Script / hourly schedule are live. Not part of the recurring
15-minute dashboard refresh -- run by hand as needed.

Usage: python3 backfill_temp_flood.py /path/to/custom-report-*.xlsx
"""
import sys
from datetime import datetime, timezone
import openpyxl
from google.cloud import bigquery

PROJECT = "apa-data-410213"
DATASET = "shelterluv"
TABLE = "TempFloodJoslyn"

COLUMN_MAP = {
    "Animal ID": "AnimalID",
    "Name": "AnimalName",
    "Species": "Species",
    "Current Status": "CurrentStatus",
    "Current Location": "CurrentLocation",
    "Location At Intake": "LocationAtIntake",
    "Intake Date": "IntakeDate",
    "Intake Time": "IntakeTime",
    "Intake Type": "IntakeType",
    "Intake Subtype": "IntakeSubtype",
    "Intake Transfer From": "IntakeTransferFrom",
    "Intake Original Source": "IntakeOriginalSource",
    "Intake From Person Name": "IntakeFromPersonName",
    "Intake From City": "IntakeFromCity",
    "Intake From County": "IntakeFromCounty",
    "Intake From Zip": "IntakeFromZip",
    "Intake Found City": "IntakeFoundCity",
    "Intake Found County": "IntakeFoundCounty",
    "Outcome Date": "OutcomeDate",
    "Outcome Type": "OutcomeType",
    "Outcome Subtype": "OutcomeSubtype",
    "Transfer To": "TransferTo",
}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "—", "-"):
        return None
    return s


def run(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = {}
    for i, h in enumerate(headers):
        if h in COLUMN_MAP:
            col_idx[COLUMN_MAP[h]] = i
    missing = set(COLUMN_MAP.values()) - set(col_idx.keys())
    if missing:
        print(f"Note: report is missing columns {missing}, will be left null.")

    ingested_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    subject = f"Manual backfill from {path.split('/')[-1]}"

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {"IngestedAt": ingested_at, "SourceEmailSubject": subject}
        for field in COLUMN_MAP.values():
            if field in col_idx:
                val = clean(raw[col_idx[field]])
                if field == "AnimalID" and val:
                    val = val.replace("APA-A-", "")
                row[field] = val
            else:
                row[field] = None
        if row.get("AnimalID"):
            rows.append(row)

    print(f"Parsed {len(rows)} rows.")

    client = bigquery.Client(project=PROJECT)
    table_ref = f"{PROJECT}.{DATASET}.{TABLE}"
    errors = client.insert_rows_json(table_ref, rows, row_ids=[None] * len(rows))
    if errors:
        print("Errors during insert:")
        for e in errors[:5]:
            print(" ", e)
    else:
        print(f"Loaded {len(rows)} rows into {table_ref}.")


if __name__ == "__main__":
    run(sys.argv[1])
